import json
import os
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


JST = timezone(timedelta(hours=9))


def load_json(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def safe_filename(name: str) -> str:
    # Windows / macOS / Linux で安全なファイル名に寄せる
    s = re.sub(r'[\\/:*?"<>|]+', "_", name)
    s = s.strip().strip(".")
    if not s:
        s = "sheet"
    return s


def cell_to_text(v: Any) -> str:
    if v is None:
        return ""
    # 日付などはopenpyxlがdatetimeで返す場合がある
    if hasattr(v, "isoformat"):
        try:
            return v.isoformat()
        except Exception:
            pass
    return str(v)


def normalize_for_search(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip().lower()


@dataclass
class HeaderSpec:
    row_start: int
    row_end: int
    join_with: str = " / "


@dataclass
class SheetSpec:
    header: HeaderSpec
    data_start_row: Optional[int]
    col_start: int
    col_end: Optional[int]
    trim_empty_rows: bool
    trim_empty_cols: bool
    max_rows_per_sheet: Optional[int]


def merge_dict(base: Dict[str, Any], override: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(base)
    for k, v in override.items():
        if isinstance(v, dict) and isinstance(out.get(k), dict):
            out[k] = merge_dict(out[k], v)
        else:
            out[k] = v
    return out


def resolve_sheet_spec(cfg: Dict[str, Any], sheet_name: str) -> SheetSpec:
    default = cfg.get("default", {})

    # 1) sheetsの完全一致
    spec_obj: Dict[str, Any] = {}
    if "sheets" in cfg and sheet_name in cfg["sheets"]:
        spec_obj = cfg["sheets"][sheet_name]
    else:
        # 2) sheet_rulesの正規表現
        for rule in cfg.get("sheet_rules", []):
            pat = rule.get("match")
            if pat and re.search(pat, sheet_name):
                spec_obj = rule
                break

    merged = merge_dict(default, spec_obj)

    header_obj = merged.get("header", {})
    # 旧形式 header_row にも対応（念のため）
    if "header_row" in merged and (not header_obj or "row_start" not in header_obj):
        hr = int(merged["header_row"])
        header_obj = {"row_start": hr, "row_end": hr, "join_with": " / "}

    header = HeaderSpec(
        row_start=int(header_obj.get("row_start", 1)),
        row_end=int(header_obj.get("row_end", header_obj.get("row_start", 1))),
        join_with=str(header_obj.get("join_with", " / ")),
    )

    data_start_row = merged.get("data_start_row", None)
    if data_start_row is not None:
        data_start_row = int(data_start_row)

    col_start = int(merged.get("col_start", 1))
    col_end = merged.get("col_end", None)
    if col_end is not None:
        col_end = int(col_end)

    trim_empty_rows = bool(merged.get("trim_empty_rows", True))
    trim_empty_cols = bool(merged.get("trim_empty_cols", False))

    max_rows_per_sheet = merged.get("max_rows_per_sheet", None)
    if max_rows_per_sheet is not None:
        max_rows_per_sheet = int(max_rows_per_sheet)

    return SheetSpec(
        header=header,
        data_start_row=data_start_row,
        col_start=col_start,
        col_end=col_end,
        trim_empty_rows=trim_empty_rows,
        trim_empty_cols=trim_empty_cols,
        max_rows_per_sheet=max_rows_per_sheet,
    )


def build_columns(ws, spec: SheetSpec, max_col: int) -> List[str]:
    col_start = spec.col_start
    col_end = spec.col_end or max_col
    col_end = min(col_end, max_col)

    headers_by_col: List[List[str]] = [[] for _ in range(col_end - col_start + 1)]

    for r in range(spec.header.row_start, spec.header.row_end + 1):
        for idx, c in enumerate(range(col_start, col_end + 1)):
            v = cell_to_text(ws.cell(row=r, column=c).value)
            v = v.strip()
            if v:
                headers_by_col[idx].append(v)

    columns: List[str] = []
    for parts in headers_by_col:
        if parts:
            columns.append(spec.header.join_with.join(parts))
        else:
            columns.append("")  # 後で補完
    # 空列名を補完
    for i, name in enumerate(columns):
        if not name:
            columns[i] = f"Column_{i+1}"
    return columns


def row_is_all_empty(row: List[str]) -> bool:
    for x in row:
        if x.strip() != "":
            return False
    return True


def trim_trailing_empty_cols(rows: List[List[str]], columns: List[str]) -> Tuple[List[List[str]], List[str]]:
    if not rows:
        return rows, columns
    last_non_empty = -1
    for r in rows:
        for i, v in enumerate(r):
            if v.strip() != "":
                last_non_empty = max(last_non_empty, i)
    if last_non_empty < 0:
        # 全部空
        return [], columns[:1] if columns else ["Column_1"]
    keep = last_non_empty + 1
    rows2 = [r[:keep] for r in rows]
    cols2 = columns[:keep]
    return rows2, cols2


def convert_one_sheet(ws, spec: SheetSpec) -> Dict[str, Any]:
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    columns = build_columns(ws, spec, max_col)

    data_start = spec.data_start_row if spec.data_start_row is not None else (spec.header.row_end + 1)
    data_start = max(1, data_start)

    col_start = spec.col_start
    col_end = spec.col_end or max_col
    col_end = min(col_end, max_col)

    rows: List[List[str]] = []
    # read_onlyでも cell() は使えるが遅くなるので、iter_rowsを使う
    # values_only=True で値だけ取り出す
    iter_min_col = col_start
    iter_max_col = col_end
    iter_min_row = data_start
    iter_max_row = max_row

    count = 0
    for row_vals in ws.iter_rows(
        min_row=iter_min_row,
        max_row=iter_max_row,
        min_col=iter_min_col,
        max_col=iter_max_col,
        values_only=True,
    ):
        row = [cell_to_text(v) for v in row_vals]
        if spec.trim_empty_rows and row_is_all_empty(row):
            continue
        rows.append(row)
        count += 1
        if spec.max_rows_per_sheet is not None and count >= spec.max_rows_per_sheet:
            break

    if spec.trim_empty_cols:
        rows, columns = trim_trailing_empty_cols(rows, columns)

    # 検索用結合テキスト
    rowText = [normalize_for_search(" ".join(r)) for r in rows]

    return {
        "columns": columns,
        "rows": rows,
        "rowText": rowText,
        "meta": {
            "header_rows": [spec.header.row_start, spec.header.row_end],
            "data_start_row": data_start,
            "col_range": [col_start, col_end],
            "trim_empty_rows": spec.trim_empty_rows,
            "trim_empty_cols": spec.trim_empty_cols,
            "rows": len(rows),
            "cols": len(columns),
        },
    }


def main() -> int:
    config_path = sys.argv[1] if len(sys.argv) >= 2 else "config.json"
    cfg = load_json(config_path)

    workbook = cfg.get("workbook")
    if not workbook:
        print("config.json に workbook を指定してください。例: \"workbook\": \"input.xlsx\"")
        return 2

    out_dir = Path(cfg.get("output_dir", "viewer/data"))
    sheets_dir = out_dir / "sheets"
    ensure_dir(sheets_dir)

    skip_hidden = bool(cfg.get("default", {}).get("skip_hidden_sheets", True))

    # data_only=True: 数式は「計算済みの値」を読み取る（未計算だとNoneになることあり）
    # read_only=True: メモリ節約
    wb = load_workbook(workbook, data_only=True, read_only=True)

    index_items: List[Dict[str, Any]] = []
    generated_at = datetime.now(JST).isoformat()

    for ws in wb.worksheets:
        if skip_hidden and ws.sheet_state != "visible":
            continue

        spec = resolve_sheet_spec(cfg, ws.title)
        sheet_json = convert_one_sheet(ws, spec)

        file_base = safe_filename(ws.title)
        # 同名が出た場合の衝突回避
        filename = f"{file_base}.json"
        i = 2
        while (sheets_dir / filename).exists():
            filename = f"{file_base}_{i}.json"
            i += 1

        sheet_json_out = {
            "sheet": ws.title,
            "generated_at": generated_at,
            **sheet_json,
        }

        with open(sheets_dir / filename, "w", encoding="utf-8") as f:
            json.dump(sheet_json_out, f, ensure_ascii=False)

        index_items.append({
            "sheet": ws.title,
            "file": f"sheets/{filename}",
            "rows": sheet_json["meta"]["rows"],
            "cols": sheet_json["meta"]["cols"],
            "header_rows": sheet_json["meta"]["header_rows"],
            "data_start_row": sheet_json["meta"]["data_start_row"],
        })

        print(f"[OK] {ws.title} -> {filename}  rows={sheet_json['meta']['rows']} cols={sheet_json['meta']['cols']}")

    index = {
        "generated_at": generated_at,
        "workbook": str(workbook),
        "sheets": index_items,
    }

    with open(out_dir / "index.json", "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False)

    print(f"\n[DONE] index.json を生成しました: {out_dir / 'index.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
